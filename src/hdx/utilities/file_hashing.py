import hashlib
import logging
import zipfile
from io import BytesIO, IOBase
from os import fstat
from typing import Tuple

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from hdx.utilities.zip_crc import get_crc_sum, get_zip_crcs_buffer, get_zip_crcs_fp

logger = logging.getLogger(__name__)

zip_signature = b"PK\x03\x04"


def hash_excel_buffer(buffer: bytes) -> str:
    """Hash the sheets in an Excel XLSX file given in a buffer using MD5

    Args:
        buffer (bytes): Excel XLSX file buffer

    Returns:
        str: MD5 hash of the sheets
    """
    file_stream = BytesIO(buffer)
    md5hash = hashlib.md5()
    try:
        workbook = load_workbook(filename=file_stream, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            md5hash.update(sheet_name.encode("utf-8"))

            for row in sheet.iter_rows(values_only=True):
                md5hash.update(str(row).encode("utf-8"))
        return md5hash.hexdigest()
    except (zipfile.BadZipFile, InvalidFileException):
        logger.error("Error hashing xlsx: The provided data is not a valid .xlsx file.")
        return ""
    except (ValueError, IndexError, AttributeError) as e:
        logger.error(
            f"Error hashing xlsx: The Excel file is corrupted or unreadable: {e}"
        )
        return ""
    except Exception as e:
        logger.error(f"Unexpected error hashing xlsx: {e}")
        return ""
    finally:
        if "workbook" in locals():
            workbook.close()
        file_stream.close()


def hash_excel_fp(fp: IOBase) -> str:
    """Hash the sheets in an Excel XLSX file given as a file pointer using MD5

    Args:
        fp (IOBase): Excel file pointer

    Returns:
        str: MD5 hash of the sheets
    """
    return hash_excel_buffer(fp.read())


def crc_zip_buffer(buffer: bytes) -> str:
    """Get sum of CRC32s for all files in a zip given a buffer

    Args:
         buffer (bytes): Zip in buffer

    Returns:
        str: Sum of the CRC32
    """

    file_crcs = get_zip_crcs_buffer(buffer)
    return get_crc_sum(file_crcs)


def crc_zip_fp(fp: IOBase) -> str:
    """Get sum of CRC32s for all files in a zip given a file pointer

    Args:
        fp (IOBase): Zip file pointer

    Returns:
        str: Sum of the CRC32
    """
    file_crcs = get_zip_crcs_fp(fp)
    return get_crc_sum(file_crcs)


def get_size_and_hash(filepath: str, file_format: str) -> Tuple[int, str]:
    """Return the size and hash of file

    Args:
        filepath: Path to file
        file_format (str): File format

    Returns:
        Tuple[int, str]: Tuple (size, hash)
    """
    with open(filepath, "rb") as fp:
        size = fstat(fp.fileno()).st_size
        signature = fp.read(4)
        if signature == zip_signature:  # zip, xlsx etc.
            if file_format.lower() == "xlsx":
                buffer = bytearray(signature)
                while chunk := fp.read(4096):
                    buffer.extend(chunk)
                hashval = hash_excel_buffer(buffer)
                if not hashval:
                    hashval = crc_zip_buffer(buffer)
                    if not hashval:
                        hashval = hashlib.md5(buffer).hexdigest()  # fallback
                del buffer
                return size, hashval
            else:
                crc_sum = crc_zip_fp(fp)
                if crc_sum:
                    return size, crc_sum
                fp.seek(4)
        md5hash = hashlib.md5(signature)
        while chunk := fp.read(4096):
            md5hash.update(chunk)
        return size, md5hash.hexdigest()
